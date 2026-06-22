"""
Génère ankorstore_import.xlsx : fichier d'import Ankorstore (feuille "Vos
produits") pour les SKUs manquants sur Ankorstore, à partir des fournisseurs
cibles (BAVOUX, DELORME, Navarro, NPC, VEINIERE).

Source des données (Supabase) :
  - skus (statut=visible, stock>0)               → SKUs vendables + stock
  - produits (nom, prix_vente_ht, image_url,
    fournisseur, id_wizi)                         → fiche produit
  - produits_ankorstore_variants (sku)            → ce qui existe déjà sur Ankorstore
  - ankorstore_ignores (sku)                      → exclusions manuelles déjà décidées

Les variations sont regroupées par id_wizi (même produit Wizishop). Pour
chaque groupe, la description BtoB est générée une seule fois (1er SKU du
groupe) via l'API Anthropic (ANTHROPIC_API_KEY).

Affiche un résumé (produits/variations/avertissements) et demande confirmation
avant tout appel à l'API Anthropic et avant l'écriture du fichier.

Usage:
    python generate_ankorstore_xlsx.py              # résumé puis confirmation
    python generate_ankorstore_xlsx.py --dry-run     # résumé seul, aucun appel Anthropic
    python generate_ankorstore_xlsx.py --yes         # saute la confirmation interactive
    python generate_ankorstore_xlsx.py --no-description  # sans appel Anthropic, description
                                                           # générique par défaut (à affiner
                                                           # manuellement après import)

Credentials lus depuis .env : SUPABASE_URL, SUPABASE_KEY, ANTHROPIC_API_KEY (sauf avec
--no-description, qui ne nécessite pas ANTHROPIC_API_KEY)
"""

import argparse
import os
import re
import sys
import time
from collections import defaultdict
from contextlib import contextmanager

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

load_dotenv()

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "ankorstore_template.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "ankorstore_import.xlsx")

# ── Mock Streamlit (identique aux autres scripts du projet) ───────────────────

class _Secrets:
    def __getitem__(self, key):
        val = os.environ.get(key)
        if val is None:
            raise KeyError(f"Secret manquant : {key!r} — vérifie ton .env")
        return val
    def get(self, key, default=None):
        return os.environ.get(key, default)
    def __contains__(self, key):
        return key in os.environ


class _SessionState(dict):
    pass


@contextmanager
def _noop_spinner(text=""):
    yield


class _MockSt:
    secrets       = _Secrets()
    session_state = _SessionState()

    def warning(self, msg):      print(f"    ⚠️  {msg}")
    def error(self, msg):        print(f"    ❌ {msg}")
    def success(self, msg):      print(f"    ✅ {msg}")
    def info(self, msg):         print(f"    ℹ️  {msg}")
    def write(self, *a, **kw):   pass
    def subheader(self, *a):     pass
    def spinner(self, text=""):  return _noop_spinner(text)

    def cache_data(self, **kwargs):
        def decorator(fn): return fn
        return decorator


sys.modules["streamlit"] = _MockSt()  # type: ignore

from supabase_api import select  # noqa: E402

# ── Constantes métier ──────────────────────────────────────────────────────────

FOURNISSEURS_CIBLES = {"BAVOUX", "DELORME", "Navarro", "NPC", "VEINIERE"}

PRIX_GROSSISTE_COEF  = 2.5
TVA_PCT              = 20
UNITES_PAR_PAQUET    = 1
FABRIQUE_EN          = "FR"
COMPOSITION          = "Acétate de cellulose"

ANTHROPIC_MODEL = "claude-sonnet-4-6"
DESCRIPTION_MIN_LEN_DUR    = 30   # minimum imposé par Ankorstore (champ obligatoire)
DESCRIPTION_MIN_LEN_SOUPLE = 500  # minimum demandé par la consigne (avertissement seulement)

DESCRIPTION_PROMPT_TEMPLATE = (
    "Tu rédiges une description produit BtoB pour Ankorstore en français pour "
    "Pique&Pince, marque française d'accessoires cheveux faits main en acétate "
    "de cellulose. \nProduit : {nom}. \nRègles : 500 caractères minimum, liste "
    "à puces d'au moins 5 éléments, ton professionnel BtoB pour revendeurs, "
    "mentionner : fait main en France, acétate de cellulose bio-sourcé, pochette "
    "velours incluse, fabriqué en France. Pas de saut de ligne. Pas du mot "
    "artisanal."
)

# Description par défaut utilisée en mode --no-description (la colonne est
# obligatoire côté Ankorstore — on ne peut pas la laisser vide).
DEFAULT_DESCRIPTION = (
    "Accessoire cheveux fait main en France en acétate de cellulose bio-sourcé. "
    "Livré dans une pochette velours Pique & Pince. Matière : acétate de "
    "cellulose. Fabriqué en France."
)

# ── Colonnes du template "Vos produits" ────────────────────────────────────────
# NB : la consigne d'origine indiquait Col AE pour "Sans cruauté", mais dans le
# fichier products-3.xlsx réel, AE = "Meilleure vente" et AG = "Sans cruauté".
# On utilise donc AG pour rester fidèle au format réel du template (cf. résumé
# affiché en cas d'écart) — un X en AE aurait marqué tous les produits comme
# meilleures ventes, ce qui n'est ni demandé ni souhaitable.

COLUMNS = {
    "sku":               "A",
    "nom":               "B",
    "description":       "C",
    "couleur":           "E",
    "image_variante":    "G",
    "image_1":           "H",
    "prix_gros":         "M",
    "prix_detail":       "N",
    "tva":               "O",
    "unites_par_paquet": "Q",
    "stock":             "R",
    "fabrique_en":       "S",
    "composition":       "Y",
    "sans_cruaute":      "AG",
    "fait_main":         "AI",
}

# Mots-clés attendus dans l'en-tête (ligne 1) de chaque colonne, pour détecter
# tout décalage si le template change un jour.
EXPECTED_HEADER_KEYWORDS = {
    "A": "SKU", "B": "Nom du produit", "C": "Description du produit",
    "E": "Couleurs des variants", "G": "Image de la variante", "H": "Image 1",
    "M": "Prix de gros", "N": "Prix de détail", "O": "Taux de TVA",
    "Q": "Nombre d'unités par paquet", "R": "Stock", "S": "Fabriqué en",
    "Y": "Composition", "AG": "Sans cruauté", "AI": "Fait main",
}


def _check_template_headers(ws):
    problems = []
    for letter, keyword in EXPECTED_HEADER_KEYWORDS.items():
        col_idx = column_index_from_string(letter)
        header = str(ws.cell(row=1, column=col_idx).value or "")
        if keyword.lower() not in header.lower():
            problems.append(f"  colonne {letter} : attendu {keyword!r}, "
                             f"trouvé {header.splitlines()[0]!r}")
    if problems:
        raise RuntimeError(
            "Le template ankorstore_template.xlsx ne correspond plus au mapping "
            "de colonnes attendu :\n" + "\n".join(problems))


# ── Récupération des SKUs manquants ───────────────────────────────────────────

def get_missing_skus():
    """SKUs visibles avec stock>0, chez un fournisseur cible, absents
    d'Ankorstore et non explicitement ignorés."""
    skus_rows = select("skus", "select=sku,stock&statut=eq.visible&stock=gt.0") or []

    ankorstore_rows = select("produits_ankorstore_variants", "select=sku") or []
    ankorstore_sku_set = {r["sku"] for r in ankorstore_rows if r.get("sku")}

    ignores_rows = select("ankorstore_ignores", "select=sku") or []
    ignores_set = {r["sku"] for r in ignores_rows if r.get("sku")}

    produits_rows = select("produits",
        "select=sku,nom,prix_vente_ht,image_url,fournisseur,id_wizi") or []
    prod_map = {p["sku"]: p for p in produits_rows}

    missing = []
    for s in skus_rows:
        sku = s.get("sku")
        if not sku or sku in ankorstore_sku_set or sku in ignores_set:
            continue
        prod = prod_map.get(sku)
        if not prod:
            continue
        if prod.get("fournisseur") not in FOURNISSEURS_CIBLES:
            continue
        missing.append({
            "sku":           sku,
            "stock":         int(s.get("stock") or 0),
            "nom":           prod.get("nom") or "",
            "prix_vente_ht": float(prod.get("prix_vente_ht") or 0),
            "image_url":     prod.get("image_url") or "",
            "fournisseur":   prod.get("fournisseur"),
            "id_wizi":       prod.get("id_wizi"),
        })
    return missing


def resolve_group(id_wizi, variations):
    """Détermine le nom du produit parent, son image, et le libellé de couleur
    de chaque variation, en s'appuyant sur les frères/sœurs en base (même
    id_wizi) plutôt que sur les seules variations manquantes — exactement la
    même logique que create_etsy_listings.py."""
    siblings = []
    if id_wizi:
        siblings = select("produits",
            f"select=sku,nom,prix_vente_ht,image_url&id_wizi=eq.{id_wizi}") or []

    parent_row = None
    for cand in siblings:
        cand_nom = cand.get("nom") or ""
        if not cand_nom:
            continue
        autres = [s for s in siblings if s["sku"] != cand["sku"]]
        if autres and all((s.get("nom") or "").startswith(cand_nom) for s in autres):
            parent_row = cand
            break

    if parent_row:
        parent_nom   = parent_row.get("nom") or ""
        parent_image = parent_row.get("image_url") or variations[0]["image_url"]
    elif len(variations) > 1:
        noms = [v["nom"] for v in variations]
        parent_nom   = os.path.commonprefix(noms).rstrip(" -–—") or noms[0]
        parent_image = variations[0]["image_url"]
    else:
        parent_nom   = variations[0]["nom"]
        parent_image = variations[0]["image_url"]

    out = []
    for v in variations:
        couleur = ""
        if parent_nom and v["nom"].startswith(parent_nom):
            couleur = v["nom"][len(parent_nom):].strip(" -–—")
        out.append({**v, "couleur": couleur})

    # Garantit une couleur unique et non vide par ligne quand le groupe a
    # plusieurs variations. Cas réel observé : deux SKUs (ex. PEI015CARAMEL /
    # PEI015ECLUXE) partageant exactement le même "nom" en base (doublon de
    # saisie Wizishop) — la soustraction du préfixe ne donne alors aucune
    # couleur pour aucun des deux. On retombe sur le suffixe distinctif du
    # SKU (après préfixe commun), qui est garanti unique puisque les SKUs le
    # sont.
    if len(out) > 1:
        par_couleur = defaultdict(list)
        for v in out:
            par_couleur[v["couleur"]].append(v)
        a_corriger = [v for couleur, rows in par_couleur.items()
                      for v in rows if not couleur or len(rows) > 1]
        if a_corriger:
            prefixe_sku = os.path.commonprefix([v["sku"] for v in out])
            for v in a_corriger:
                v["couleur"] = v["sku"][len(prefixe_sku):] or v["sku"]

    return parent_nom, parent_image, out


def build_plans(missing):
    groups = defaultdict(list)
    for m in missing:
        groups[m["id_wizi"]].append(m)

    plans = []
    for idw, variations in groups.items():
        variations = sorted(variations, key=lambda v: v["sku"])
        parent_nom, parent_image, variations_lbl = resolve_group(idw, variations)

        warnings = []
        if not parent_nom:
            warnings.append("Nom de produit introuvable")
        if not parent_image:
            warnings.append("Aucune image principale disponible")
        for v in variations_lbl:
            if not v["image_url"]:
                warnings.append(f"Image manquante pour la variation {v['sku']}")
            if v["prix_vente_ht"] <= 0:
                warnings.append(f"Prix de vente HT à 0 ou manquant pour {v['sku']}")

        plans.append({
            "id_wizi":      idw,
            "nom":          parent_nom,
            "image_parent": parent_image,
            "variations":   variations_lbl,
            "warnings":     warnings,
            "errors":       [] if parent_nom else ["Nom de produit introuvable — groupe ignoré"],
            "description":  None,
        })

    plans.sort(key=lambda p: p["nom"])
    return plans


# ── Résumé console ─────────────────────────────────────────────────────────────

def print_summary(plans):
    nb_err       = sum(1 for p in plans if p["errors"])
    nb_ok        = len(plans) - nb_err
    nb_variations = sum(len(p["variations"]) for p in plans)

    print("\n" + "=" * 100)
    print(f"  RÉSUMÉ — {len(plans)} produit(s) / {nb_variations} variation(s) manquant(s) sur Ankorstore")
    print("=" * 100)

    for p in plans:
        flag = "❌" if p["errors"] else ("⚠️ " if p["warnings"] else "✅")
        skus = ", ".join(v["sku"] for v in p["variations"])
        print(f"\n{flag} {p['nom'][:70]}")
        print(f"     {len(p['variations'])} variation(s) — SKUs : {skus}")
        for e in p["errors"]:
            print(f"     ❌ {e}")
        for w in p["warnings"]:
            print(f"     ⚠️  {w}")

    print("\n" + "-" * 100)
    print(f"  Total : {nb_ok} produit(s) prêt(s) — {nb_err} en erreur (seront ignorés)")
    print(f"  Soit {nb_variations} ligne(s) au total dans le fichier Excel")
    print("-" * 100)


# ── Génération de description via Anthropic ───────────────────────────────────

def generate_description(nom):
    prompt = DESCRIPTION_PROMPT_TEMPLATE.format(nom=nom)
    r = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key":         os.environ["ANTHROPIC_API_KEY"],
            "anthropic-version": "2023-06-01",
            "content-type":      "application/json",
        },
        json={
            "model":      ANTHROPIC_MODEL,
            "max_tokens": 800,
            "messages":   [{"role": "user", "content": prompt}],
        },
        timeout=60,
    )
    r.raise_for_status()
    texte = r.json()["content"][0]["text"].strip()
    texte = re.sub(r"\s*\n+\s*", " ", texte)   # "Pas de saut de ligne"
    texte = re.sub(r" {2,}", " ", texte).strip()
    return texte


# ── Écriture du fichier Excel ──────────────────────────────────────────────────

def write_excel(plans, template_path, output_path):
    wb = load_workbook(template_path)
    ws = wb["Vos produits"]
    _check_template_headers(ws)
    col = {k: column_index_from_string(v) for k, v in COLUMNS.items()}

    row_idx = 2
    for p in plans:
        for i, v in enumerate(p["variations"]):
            ws.cell(row=row_idx, column=col["sku"], value=v["sku"])
            ws.cell(row=row_idx, column=col["nom"], value=p["nom"])
            if i == 0:
                ws.cell(row=row_idx, column=col["description"], value=p["description"])
                ws.cell(row=row_idx, column=col["image_1"], value=p["image_parent"])
            ws.cell(row=row_idx, column=col["couleur"], value=v["couleur"])
            ws.cell(row=row_idx, column=col["image_variante"], value=v["image_url"])
            ws.cell(row=row_idx, column=col["prix_gros"],
                    value=round(v["prix_vente_ht"] / PRIX_GROSSISTE_COEF, 2))
            ws.cell(row=row_idx, column=col["prix_detail"],
                    value=round(v["prix_vente_ht"] * 1.2, 2))
            ws.cell(row=row_idx, column=col["tva"], value=TVA_PCT)
            ws.cell(row=row_idx, column=col["unites_par_paquet"], value=UNITES_PAR_PAQUET)
            ws.cell(row=row_idx, column=col["stock"], value=v["stock"])
            ws.cell(row=row_idx, column=col["fabrique_en"], value=FABRIQUE_EN)
            ws.cell(row=row_idx, column=col["composition"], value=COMPOSITION)
            ws.cell(row=row_idx, column=col["sans_cruaute"], value="X")
            ws.cell(row=row_idx, column=col["fait_main"], value="X")
            row_idx += 1

    wb.save(output_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Génère le fichier d'import Ankorstore pour les SKUs manquants")
    parser.add_argument("--dry-run", action="store_true",
                         help="Résumé uniquement, aucun appel Anthropic ni fichier généré")
    parser.add_argument("--yes", action="store_true",
                         help="Ne pas demander de confirmation interactive")
    parser.add_argument("--no-description", action="store_true",
                         help="Génère le fichier sans appeler l'API Anthropic — utilise une "
                              "description générique par défaut (à affiner manuellement)")
    args = parser.parse_args()

    if not os.path.exists(TEMPLATE_PATH):
        print(f"❌ Template introuvable : {TEMPLATE_PATH}")
        sys.exit(1)

    print("🔎 Récupération des SKUs manquants sur Ankorstore (Supabase)…")
    missing = get_missing_skus()
    print(f"   {len(missing)} SKU(s) manquant(s) chez {', '.join(sorted(FOURNISSEURS_CIBLES))}")

    if not missing:
        print("\nAucun SKU manquant — rien à générer.")
        return

    plans = build_plans(missing)
    print_summary(plans)

    ready   = [p for p in plans if not p["errors"]]
    skipped = [p for p in plans if p["errors"]]

    if args.dry_run:
        print("\n(dry-run) Aucun appel à l'API Anthropic, aucun fichier généré.")
        return

    if not ready:
        print("\nAucun produit prêt à être généré.")
        return

    if not args.no_description and "ANTHROPIC_API_KEY" not in os.environ:
        print("❌ ANTHROPIC_API_KEY absente du .env — impossible de générer les descriptions.\n"
              "   (utilise --no-description pour générer le fichier sans description)")
        sys.exit(1)

    if not args.yes:
        if args.no_description:
            question = (f"\nGénérer {len(ready)} produit(s) sans description (colonne vide) "
                         f"dans {OUTPUT_PATH!r} ? (oui/non) : ")
        else:
            question = (f"\nGénérer la description de {len(ready)} produit(s) via Claude puis "
                         f"écrire {OUTPUT_PATH!r} ? (oui/non) : ")
        reponse = input(question).strip().lower()
        if reponse not in ("oui", "o", "yes", "y"):
            print("Annulé.")
            return

    final_plans = []

    if args.no_description:
        print(f"\n📄 Génération de {len(ready)} produit(s) avec la description "
              f"générique par défaut (--no-description)…\n")
        for p in ready:
            p["description"] = DEFAULT_DESCRIPTION
            final_plans.append(p)
    else:
        print(f"\n🚀 Génération des descriptions pour {len(ready)} produit(s)…\n")
        for p in ready:
            print(f"→ {p['nom'][:60]} …", end=" ")
            try:
                description = generate_description(p["nom"])
            except Exception as e:
                print(f"❌ échec génération : {e}")
                skipped.append(p)
                continue

            if len(description) < DESCRIPTION_MIN_LEN_DUR:
                print(f"❌ description trop courte ({len(description)} car., minimum Ankorstore "
                      f"{DESCRIPTION_MIN_LEN_DUR})")
                skipped.append(p)
                continue

            if len(description) < DESCRIPTION_MIN_LEN_SOUPLE:
                print(f"⚠️  description courte ({len(description)} car., {DESCRIPTION_MIN_LEN_SOUPLE} demandés)")
            else:
                print(f"✅ {len(description)} car.")

            p["description"] = description
            final_plans.append(p)
            time.sleep(0.3)

    if not final_plans:
        print("\nAucune description générée avec succès — fichier non créé.")
        return

    write_excel(final_plans, TEMPLATE_PATH, OUTPUT_PATH)

    nb_variations = sum(len(p["variations"]) for p in final_plans)
    print(f"\n{'=' * 100}")
    print(f"  TERMINÉ — {OUTPUT_PATH}")
    print(f"  {len(final_plans)} produit(s), {nb_variations} ligne(s) écrite(s)")
    print(f"  {len(skipped)} produit(s) ignoré(s) (erreurs de données ou génération échouée)")
    print(f"{'=' * 100}")
    for p in skipped:
        raison = "; ".join(p["errors"]) if p["errors"] else "génération de description échouée"
        print(f"  ❌ {p['nom'][:60] or p['id_wizi']} : {raison}")


if __name__ == "__main__":
    main()
